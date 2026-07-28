import json
import os
from datetime import datetime

import pandas as pd

from config import config

logger = __import__('logging').getLogger('trader_system')


def _analyze_performance_dir(perf_dir, type_name, since_timestamp=None):
    """分析单个策略表现目录

    Args:
        since_timestamp: 可选，只统计此时间戳之后修改的文件
    """
    if not os.path.exists(perf_dir):
        return {'type': type_name, 'count': 0, 'files': [], 'issues': []}

    results = {'type': type_name, 'count': 0, 'files': [], 'issues': []}
    perf_dir = str(perf_dir)

    for f in os.listdir(perf_dir):
        if not f.endswith('_strategy_performance_test.csv'):
            continue
        filepath = os.path.join(perf_dir, f)
        try:
            # 如果指定了时间戳，只统计该时间戳之后修改的文件
            if since_timestamp is not None:
                mtime = os.path.getmtime(filepath)
                if mtime < since_timestamp:
                    continue
            df = pd.read_csv(filepath)
            if df.empty:
                # 只有表头无数据行：属于"无信号"的正常情况，不算问题，计入空文件统计
                results.setdefault('empty_files', []).append(f)
            else:
                results['files'].append({
                    'name': f,
                    'rows': len(df),
                    'size': os.path.getsize(filepath),
                    'mtime': os.path.getmtime(filepath),
                })
                results['count'] += len(df)
        except Exception as e:
            results['issues'].append(f'{f}: 读取失败 - {str(e)[:100]}')

    return results


def _analyze_daily_data_dir(data_dir, type_name):
    """分析每日数据目录（统计有效的 Parquet/CSV 日线文件，避免重复计数）。"""
    if not os.path.exists(data_dir):
        return {'type': type_name, 'count': 0, 'issues': []}

    results = {'type': type_name, 'count': 0, 'issues': []}
    counted = set()

    for f in os.listdir(data_dir):
        if not (f.endswith('_daily.parquet') or f.endswith('_daily.csv')):
            continue
        # 同一 code 的 Parquet 和 CSV 只计一次
        code_key = f.replace('_daily.parquet', '').replace('_daily.csv', '')
        if code_key in counted:
            continue
        counted.add(code_key)
        results['count'] += 1

    return results


def generate_quality_report(since_timestamp=None):
    """生成数据质量报告（JSON + Excel）

    Args:
        since_timestamp: 可选，只统计此时间戳之后修改的文件（用于显示本次更新明细）
    """
    report = {
        'generated_at': datetime.now().timestamp(),
        'generated_at_str': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'performance': [],
        'daily_data': [],
        'summary': {},
    }

    report['performance'].append(_analyze_performance_dir(config.STOCK_PERFORMANCE_DIR, 'A股', since_timestamp))
    report['performance'].append(_analyze_performance_dir(config.INDEX_PERFORMANCE_DIR, '指数', since_timestamp))
    report['performance'].append(_analyze_performance_dir(config.ETF_PERFORMANCE_DIR, 'ETF', since_timestamp))

    report['daily_data'].append(_analyze_daily_data_dir(config.DAILY_TRACKING_A_DIR, 'A股'))
    report['daily_data'].append(_analyze_daily_data_dir(config.DAILY_TRACKING_INDEX_DIR, '指数'))
    report['daily_data'].append(_analyze_daily_data_dir(config.DAILY_TRACKING_ETF_DIR, 'ETF'))

    total_perf = sum(p['count'] for p in report['performance'])
    total_issues = sum(len(p['issues']) for p in report['performance'])
    total_daily = sum(d['count'] for d in report['daily_data'])
    total_empty = sum(len(p.get('empty_files', [])) for p in report['performance'])

    report['summary'] = {
        'total_performance_records': total_perf,
        'total_performance_issues': total_issues,
        'total_daily_files': total_daily,
        'total_empty_files': total_empty,
        'overall_status': 'good' if total_issues == 0 else 'warning' if total_issues < 10 else 'error',
    }

    with open(config.SIGNAL_UPDATE_QUALITY_FILE, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    # 同步生成 Excel
    _write_quality_xlsx(report)

    logger.info(f'[数据质量] 生成报告: {total_perf}条记录, {total_issues}个问题')

    return report


def _write_quality_xlsx(report):
    """将报告数据写入 Excel，含多 sheet"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    xlsx_path = config.SIGNAL_UPDATE_DIR / 'quality_report.xlsx'
    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def _write_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

    def _write_data(ws, data, start_row=2):
        for r_idx, row_data in enumerate(data, start=start_row):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                cell.alignment = align_center

    # --- Sheet1: 概览 ---
    ws1 = wb.active
    ws1.title = '概览'
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 35
    ws1.column_dimensions['C'].width = 18

    overview_data = [
        ['生成时间', report['generated_at_str']],
        ['策略表现总记录数', report['summary']['total_performance_records']],
        ['数据问题数', report['summary']['total_performance_issues']],
        ['每日数据文件数', report['summary']['total_daily_files']],
        ['整体状态', report['summary']['overall_status']],
    ]
    for r_idx, (k, v) in enumerate(overview_data, 1):
        ws1.cell(row=r_idx, column=1, value=k).font = Font(bold=True)
        ws1.cell(row=r_idx, column=1).border = thin_border
        ws1.cell(row=r_idx, column=2, value=v).border = thin_border
        ws1.cell(row=r_idx, column=2).alignment = align_left

    # --- Sheet2: 策略表现明细 ---
    ws2 = wb.create_sheet('策略表现明细')
    headers2 = ['市场类型', '文件数', '总记录数', '问题数']
    _write_header(ws2, headers2)
    perf_rows = [
        [p['type'], len(p['files']), p['count'], len(p['issues'])]
        for p in report['performance']
    ]
    _write_data(ws2, perf_rows)
    for col in ['A', 'B', 'C', 'D']:
        ws2.column_dimensions[col].width = 18

    # --- Sheet3: 每日数据明细 ---
    ws3 = wb.create_sheet('每日数据明细')
    headers3 = ['市场类型', '每日数据文件数', '问题数']
    _write_header(ws3, headers3)
    daily_rows = [
        [d['type'], d['count'], len(d['issues'])]
        for d in report['daily_data']
    ]
    _write_data(ws3, daily_rows)
    for col in ['A', 'B', 'C']:
        ws3.column_dimensions[col].width = 20

    # --- Sheet4: 数据问题 ---
    ws4 = wb.create_sheet('数据问题')
    headers4 = ['市场类型', '问题描述']
    _write_header(ws4, headers4)
    issue_rows = []
    for p in report['performance']:
        for issue in p['issues']:
            issue_rows.append([p['type'], issue])
    for d in report['daily_data']:
        for issue in d['issues']:
            issue_rows.append([d['type'], issue])
    _write_data(ws4, issue_rows)
    ws4.column_dimensions['A'].width = 15
    ws4.column_dimensions['B'].width = 60

    wb.save(str(xlsx_path))
    return xlsx_path


def get_quality_report():
    """获取最新数据质量报告"""
    path = config.SIGNAL_UPDATE_QUALITY_FILE
    if not path.exists():
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def get_performance_stats(code=None):
    """获取策略表现统计"""
    stats = []
    dirs = [
        (config.STOCK_PERFORMANCE_DIR, 'A股'),
        (config.INDEX_PERFORMANCE_DIR, '指数'),
        (config.ETF_PERFORMANCE_DIR, 'ETF'),
    ]

    for perf_dir, type_name in dirs:
        if not os.path.exists(str(perf_dir)):
            continue
        for f in os.listdir(str(perf_dir)):
            if not f.endswith('_strategy_performance_test.csv'):
                continue
            if code and not f.startswith(code):
                continue
            filepath = os.path.join(str(perf_dir), f)
            try:
                df = pd.read_csv(filepath)
                if not df.empty:
                    avg_win_rate = df['胜率(%)'].mean() if '胜率(%)' in df.columns else 0
                    avg_return = df['简易收益率(%)'].mean() if '简易收益率(%)' in df.columns else 0
                    stats.append({
                        'code': f.replace('_strategy_performance_test.csv', '').replace('_buy', '').replace('_sell', ''),
                        'type': f.split('_')[-2],
                        'market': type_name,
                        'total_trades': df['交易次数'].sum() if '交易次数' in df.columns else 0,
                        'avg_win_rate': round(avg_win_rate, 2),
                        'avg_return': round(avg_return, 2),
                        'file': f,
                    })
            except Exception:
                pass

    return stats
